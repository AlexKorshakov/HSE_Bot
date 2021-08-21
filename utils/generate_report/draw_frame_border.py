
from openpyxl.styles import Border, Side


async def set_border(ws):
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    columns = ws.max_column
    rows = ws.max_row

    for col in range(1, columns):
        for row in range(1, rows):
            ws.cell(row=row, column=col).border = thin_border
