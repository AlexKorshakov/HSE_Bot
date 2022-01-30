import datetime
import os
from math import ceil

from loguru import logger
from openpyxl.drawing.image import Image

from data.category import CATEGORY_LIST, ELIMINATION_TIME
from utils.generate_report.mip_report_settings import CATEGORY_LIST_VALUES
from utils.generate_report.sheet_formatting.set_alignment import set_mip_alignment
from utils.generate_report.sheet_formatting.set_font import sets_report_font, set_report_font
from utils.generate_report.sheet_formatting.set_frame_border import set_range_border
from utils.img_processor.insert_img import image_preparation, insert_images
from utils.json_worker.read_json_file import read_json_file

not_found: str = 'не выявлено'
not_tested: str = 'не проверялось'
be_away: str = 'отсутствовал'
check_mark_true: str = 'V'
check_mark_false: str = '□'

start_photo_row: int = 52
photo_column: str = 'C'
description_column: int = 7
photo_height: int = 400


async def set_report_body_values(worksheet):
    """
    :param worksheet:
    :return:
    """
    values = [
        {"coordinate": "C2", "value": "МОСИНЖПРОЕКТ", "row": "2", "column": "3"},
        {"coordinate": "D2", "value": "Отчет о ночной смены", "row": "2", "column": "4"},
        {"coordinate": "C3", "value": "ЛО-МИП-УОТиПБ--", "row": "3", "column": "3"},
        {"coordinate": "D3", "value": "Значение", "row": "3", "column": "4"},
        {"coordinate": "F3", "value": "Примечание", "row": "3", "column": "6"},
        {"coordinate": "C4", "value": "Общая информация", "row": "4", "column": "3"},
        {"coordinate": "C5", "value": "Обход", "row": "5", "column": "3"},
        {"coordinate": "D5", "value": "Первичный", "row": "5", "column": "4"},
        {"coordinate": "C6", "value": "Дата", "row": "6", "column": "3"},
        {"coordinate": "C7", "value": "Подрядчик", "row": "7", "column": "3"},
        {"coordinate": "D7", "value": "стм.", "row": "7", "column": "4"},
        {"coordinate": "C8", "value": "Субподрядчик", "row": "8", "column": "3"},
        {"coordinate": "C9", "value": "Проект", "row": "9", "column": "3"},
        {"coordinate": "D9", "value": "", "row": "9", "column": "4"},
        {"coordinate": "C10", "value": "Комиссия", "row": "10", "column": "3"},
        {"coordinate": "D10", "value": "Функция", "row": "10", "column": "4"},
        {"coordinate": "F10", "value": "ФИО", "row": "10", "column": "6"},
        {"coordinate": "D11", "value": "□", "row": "11", "column": "4"},
        {"coordinate": "E11", "value": "Инспектирующие", "row": "11", "column": "5"},
        {"coordinate": "D12", "value": f"{check_mark_true}", "row": "12", "column": "4"},
        {"coordinate": "E12", "value": "Руководитель строительства", "row": "12", "column": "5"},
        {"coordinate": "F12", "value": f"{be_away}", "row": "12", "column": "6"},
        {"coordinate": "D13", "value": f"{check_mark_true}", "row": "13", "column": "4"},
        {"coordinate": "E13", "value": "Специалист отдела ПБ", "row": "13", "column": "5"},
        {"coordinate": "D14", "value": f"{check_mark_true}", "row": "14", "column": "4"},
        {"coordinate": "E14", "value": "Инженер СК", "row": "14", "column": "5"},
        {"coordinate": "F14", "value": f"{be_away}", "row": "14", "column": "6"},
        {"coordinate": "D15", "value": f"{check_mark_true}", "row": "15", "column": "4"},
        {"coordinate": "E15", "value": "Подрядчик", "row": "15", "column": "5"},
        {"coordinate": "F15", "value": f"{be_away}", "row": "15", "column": "6"},
        {"coordinate": "D16", "value": f"{check_mark_true}", "row": "16", "column": "4"},
        {"coordinate": "E16", "value": "Субподрядчик", "row": "16", "column": "5"},
        {"coordinate": "C17", "value": "Охрана труда, промышленная безопасность и охрана окружающей среды", "row": "17",
         "column": "3"},
        {"coordinate": "C18", "value": "Наблюдения", "row": "18", "column": "3"},
        {"coordinate": "D18", "value": "Категория несоответствия", "row": "18", "column": "4"},
        {"coordinate": "F18", "value": "№", "row": "18", "column": "6"},
        {"coordinate": "G18", "value": "Несоответствие", "row": "18", "column": "7"},
        {"coordinate": "H18", "value": "Срок", "row": "18", "column": "8"},

        {"coordinate": "D19", "value": f"{check_mark_false}", "row": "19", "column": "4"},
        {"coordinate": "E19", "value": "Документы ОТ и ПБ", "row": "19", "column": "5"},
        {"coordinate": "G19", "value": f"{not_tested}", "row": "19", "column": "7"},

        {"coordinate": "D20", "value": f"{check_mark_false}", "row": "20", "column": "4"},
        {"coordinate": "E20", "value": "Обучение/аттестация/квалификация", "row": "20", "column": "5"},
        {"coordinate": "G20", "value": f"{not_found}", "row": "20", "column": "7"},

        {"coordinate": "D21", "value": f"{check_mark_true}", "row": "21", "column": "4"},
        {"coordinate": "E21", "value": "СИЗ", "row": "21", "column": "5"},
        {"coordinate": "G21", "value": f"{not_found}", "row": "21", "column": "7"},

        {"coordinate": "D22", "value": f"{check_mark_true}", "row": "22", "column": "4"},
        {"coordinate": "E22", "value": "Механизмы и оборудование", "row": "22", "column": "5"},
        {"coordinate": "G22", "value": f"{not_found}", "row": "22", "column": "7"},

        {"coordinate": "D23", "value": f"{check_mark_true}", "row": "23", "column": "4"},
        {"coordinate": "E23", "value": "ТС/Спецтехника", "row": "23", "column": "5"},
        {"coordinate": "G23", "value": f"{not_found}", "row": "23", "column": "7"},

        {"coordinate": "D24", "value": f"{check_mark_true}", "row": "24", "column": "4"},
        {"coordinate": "E24", "value": "Знаки безопасности/ограждения", "row": "24", "column": "5"},
        {"coordinate": "G24", "value": f"{not_found}", "row": "24", "column": "7"},

        {"coordinate": "D25", "value": f"{check_mark_true}", "row": "25", "column": "4"},
        {"coordinate": "E25", "value": "Земляные работы", "row": "25", "column": "5"},
        {"coordinate": "G25", "value": f"{not_found}", "row": "25", "column": "7"},

        {"coordinate": "D26", "value": f"{check_mark_true}", "row": "26", "column": "4"},
        {"coordinate": "E26", "value": "Электробезопасность", "row": "26", "column": "5"},
        {"coordinate": "G26", "value": f"{not_found}", "row": "26", "column": "7"},

        {"coordinate": "D27", "value": f"{check_mark_true}", "row": "27", "column": "4"},
        {"coordinate": "E27", "value": "Бетонные работы", "row": "27", "column": "5"},
        {"coordinate": "G27", "value": f"{not_found}", "row": "27", "column": "7"},

        {"coordinate": "D28", "value": f"{check_mark_true}", "row": "28", "column": "4"},
        {"coordinate": "E28", "value": "ГПМ", "row": "28", "column": "5"},
        {"coordinate": "G28", "value": f"{not_found}", "row": "28", "column": "7"},

        {"coordinate": "D29", "value": f"{check_mark_true}", "row": "29", "column": "4"},
        {"coordinate": "E29", "value": "Замкнутые пространства", "row": "29", "column": "5"},
        {"coordinate": "G29", "value": f"{not_found}", "row": "29", "column": "7"},

        {"coordinate": "D30", "value": f"{check_mark_true}", "row": "30", "column": "4"},
        {"coordinate": "E30", "value": "Ручные инструменты", "row": "30", "column": "5"},
        {"coordinate": "G30", "value": f"{not_found}", "row": "30", "column": "7"},

        {"coordinate": "D31", "value": f"{check_mark_true}", "row": "31", "column": "4"},
        {"coordinate": "E31", "value": "Работы на высоте", "row": "31", "column": "5"},
        {"coordinate": "G31", "value": f"{not_found}", "row": "31", "column": "7"},

        {"coordinate": "D32", "value": f"{check_mark_true}", "row": "32", "column": "4"},
        {"coordinate": "E32", "value": "Огневые работы", "row": "32", "column": "5"},
        {"coordinate": "G32", "value": f"{not_found}", "row": "32", "column": "7"},

        {"coordinate": "D33", "value": f"{check_mark_true}", "row": "33", "column": "4"},
        {"coordinate": "E33", "value": "Оборудование под давлением", "row": "33", "column": "5"},
        {"coordinate": "G33", "value": f"{not_found}", "row": "33", "column": "7"},

        {"coordinate": "D34", "value": f"{check_mark_true}", "row": "34", "column": "4"},
        {"coordinate": "E34", "value": "Пожарная безопасность", "row": "34", "column": "5"},
        {"coordinate": "G34", "value": f"{not_found}", "row": "34", "column": "7"},

        {"coordinate": "D35", "value": f"{check_mark_true}", "row": "35", "column": "4"},
        {"coordinate": "E35", "value": "Первая помощь", "row": "35", "column": "5"},
        {"coordinate": "G35", "value": f"{not_found}", "row": "35", "column": "7"},

        {"coordinate": "D36", "value": f"{check_mark_false}", "row": "36", "column": "4"},
        {"coordinate": "E36", "value": "Химические, биологические факторы", "row": "36", "column": "5"},
        {"coordinate": "G36", "value": f"{not_tested}", "row": "36", "column": "7"},

        {"coordinate": "D37", "value": f"{check_mark_true}", "row": "37", "column": "4"},
        {"coordinate": "E37", "value": "Санитарные требования", "row": "37", "column": "5"},
        {"coordinate": "G37", "value": f"{not_found}", "row": "37", "column": "7"},

        {"coordinate": "D38", "value": f"{check_mark_true}", "row": "38", "column": "4"},
        {"coordinate": "E38", "value": "Складирование", "row": "38", "column": "5"},
        {"coordinate": "G38", "value": f"{not_found}", "row": "38", "column": "7"},

        {"coordinate": "D39", "value": f"{check_mark_true}", "row": "39", "column": "4"},
        {"coordinate": "E39", "value": "Безопасные проходы", "row": "39", "column": "5"},
        {"coordinate": "G39", "value": f"{not_found}", "row": "39", "column": "7"},

        {"coordinate": "D40", "value": f"{check_mark_true}", "row": "40", "column": "4"},
        {"coordinate": "E40", "value": "Отходы", "row": "40", "column": "5"},
        {"coordinate": "G40", "value": f"{not_found}", "row": "40", "column": "7"},

        {"coordinate": "D41", "value": f"{check_mark_true}", "row": "41", "column": "4"},
        {"coordinate": "E41", "value": "Дежурное освещение", "row": "41", "column": "5"},
        {"coordinate": "G41", "value": f"{not_found}", "row": "41", "column": "7"},

        {"coordinate": "D42", "value": f"{check_mark_true}", "row": "42", "column": "4"},
        {"coordinate": "E42", "value": "Другое", "row": "42", "column": "5"},
        {"coordinate": "G42", "value": f"{not_found}", "row": "42", "column": "7"},

        {"coordinate": "C43", "value": "Дополнительная информация", "row": "43", "column": "3"},
        {"coordinate": "C44",
         "value": 'Данное сообщение рассылается Блоком по качеству, охране труда, промышленной безопасности и охране '
                  'окружающей среды АО Мосинжпроект с целью информирования о состоянии площадки, производства и '
                  'документирования строительно-монтажных работ',
         "row": "44", "column": "3"}
    ]

    for val in values:
        try:
            worksheet.cell(row=int(val['row']), column=int(val['column'])).value = str(val['value'])

            if val["value"] == not_found:
                cell_range = [f"D{val['row']}:H{val['row']}",
                              {"color": "008000", "font_size": "14", "name": "Arial"}]
                await sets_report_font(worksheet, cell_range[0], params=cell_range[1])

        except Exception as err:
            logger.error(f"set_values {repr(err)}")
            return None


async def set_report_header_values(worksheet, registration_data, dataframe):
    """Заполнение заголовка отчета
    :param worksheet:
    :param registration_data:
    :return:
    """
    contractors = dataframe.general_contractor.tolist()
    contractors = list(set(list(contractors)))
    contractors_str = ''
    for item in contractors:
        if isinstance(item, float):
            continue
        if item and item != 'Подрядная организация':
            contractors_str = contractors_str + item + ', '

    date_now = datetime.datetime.now().strftime("%d.%m.%Y")
    date_then = datetime.datetime.now() - datetime.timedelta(days=1)
    date_then = date_then.strftime("%d.%m.%Y")

    day = (datetime.datetime.now()).strftime("%d")
    year = (datetime.datetime.now()).strftime("%Y")

    function = registration_data.get("function")
    name = registration_data.get("name")
    name_location = registration_data.get("name_location")
    phone_number = registration_data.get("phone_number")
    work_shift = registration_data.get("work_shift")

    if work_shift.lower() == "дневная смена":
        work_shift = 'дневной смены'
        custom_date = f"{date_now}"
    else:
        work_shift = 'о ночной смене'
        custom_date = f"{date_then} - {date_now}"

    values = [
        {"coordinate": "D2", "value": f"Отчет {work_shift} {custom_date}", "row": "2", "column": "4"},
        {"coordinate": "C3", "value": f"ЛО-МИП-УОТиПБ-{year}-{day}", "row": "3", "column": "3"},
        {"coordinate": "D6", "value": f"{custom_date}", "row": "6", "column": "4"},
        {"coordinate": "D7", "value": f"{contractors_str}", "row": "7", "column": "4"},
        {"coordinate": "D8", "value": "", "row": "8", "column": "4"},
        {"coordinate": "D9", "value": f"{name_location}", "row": "9", "column": "4"},
        {"coordinate": "F12", "value": f"{be_away}", "row": "12", "column": "6"},

        {"coordinate": "E13", "value": f"{function}", "row": "13", "column": "5"},
        {"coordinate": "E13", "value": f"{name} тел. +{phone_number}", "row": "13", "column": "6"},

        {"coordinate": "F14", "value": f"{be_away}", "row": "14", "column": "6"},
        {"coordinate": "F15", "value": f"{be_away}", "row": "15", "column": "6"},
        {"coordinate": "F16", "value": f"{be_away}", "row": "16", "column": "6"},
    ]

    for val in values:
        try:

            worksheet.cell(row=int(val['row']), column=int(val['column']), value=str(val['value']))

        except Exception as err:
            logger.error(f"set_user_values {repr(err)}")
            return None


async def set_report_violation_values(worksheet, dataframe):
    """ установить значения нарушения
    :param worksheet:
    :param dataframe:
    :return:
    """

    serial_number = 0
    val = []
    for category in CATEGORY_LIST:
        for item in range(1, dataframe.category.size):
            if dataframe.loc[item]["category"] != category:
                continue

            try:
                elimination_time = await get_elimination_time(dataframe, item)
                val.append({"description": dataframe.loc[item]["description"] + ' \\',
                            "elimination_time": elimination_time + ' \\'})
            except Exception as err:
                logger.error(f"{repr(err)}")
                return

        if not val:
            continue

        serial_number += 1
        for item in CATEGORY_LIST_VALUES:

            if category != item['value']:
                continue

            await set_worksheet_cell_value(worksheet, item, serial_number, val)

            val = []
            break


async def get_elimination_time(dataframe, item) -> str:
    """Получить время устранения
    :return: str
    """
    if dataframe.loc[item]["elimination_time"] != ELIMINATION_TIME[0] and \
            dataframe.loc[item]["elimination_time"] != ELIMINATION_TIME[1]:
        add_days = int(dataframe.loc[item]["elimination_time"].split(' ')[0])
        elimination_time = datetime.datetime.now() + datetime.timedelta(days=add_days)

        return str(elimination_time.strftime("%d.%m.%Y"))

    return dataframe.loc[item]["elimination_time"]


async def set_worksheet_cell_value(worksheet, item, serial_number, val):
    """
    :param worksheet:
    :param item:
    :param serial_number:
    :return:
    """
    worksheet.cell(row=int(item['row']), column=6, value=serial_number)
    worksheet.cell(row=int(item['row']), column=7, value="".join([str(i["description"]) + ' \n' for i in val]))
    worksheet.cell(row=int(item['row']), column=8, value="".join([str(i["elimination_time"]) + ' \n' for i in val]))

    cell_ranges = [
        [f"D{item['row']}:F{item['row']}", {"color": "FF0000", "font_size": "14", "name": "Arial"}],
        [f"G{item['row']}:G{item['row']}", {"color": "FF0000", "font_size": "12", "name": "Arial"}],
        [f"H{item['row']}:H{item['row']}", {"color": "FF0000", "font_size": "12", "name": "Arial"}],
    ]
    for cell_range in cell_ranges:
        await sets_report_font(worksheet, cell_range[0], params=cell_range[1])
        await set_mip_alignment(worksheet, cell_range[0], horizontal='left', vertical='center')

    worksheet.row_dimensions[int(item['row'])].height = await get_height_for_row(
        worksheet,
        int(item['row']),
        value="".join([str(i["description"]) + ' \n' for i in val]))


factor_of_font_size_to_width = {

    12: {
        "factor": 0.9,  # width / count of symbols at row
        "height": 25
    }
}


async def get_height_for_row(sheet, row_number: int, font_size: int = 12, value=None):
    """
    :param value:
    :param sheet:
    :param row_number:
    :param font_size:
    :return:
    """
    font_params = factor_of_font_size_to_width[font_size]
    row = list(sheet.rows)[row_number]
    height = font_params["height"]

    for index, cell in enumerate(row):
        if index != 6:
            continue

        if hasattr(cell, 'column_letter'):
            words_count_at_one_row = sheet.column_dimensions[cell.column_letter].width / font_params["factor"]

            lines = ceil(len(str(value)) / words_count_at_one_row)
            height = max(height, lines * font_params["height"])

            return height


async def set_photographic_materials_values(worksheet):
    """
    :param worksheet:

    """

    values = [
        {"coordinate": "C50", "value": "Фотофиксация", "row": "50", "column": "3"},
        {"coordinate": "C51", "value": "ФОТО", "row": "51", "column": "3"},
        {"coordinate": "F51", "value": "№", "row": "51", "column": "6"},
        {"coordinate": "C51", "value": "Примечание", "row": "51", "column": "7"},
    ]

    for val in values:
        try:

            worksheet.cell(row=int(val['row']), column=int(val['column']), value=str(val['value']))

        except Exception as err:
            logger.error(f"set_photographic_materials_values {repr(err)}")
            return None


async def set_photographic_materials(worksheet, violation_data: str, num_data: int):
    """Вставка фото и описания наоушения га сстраницу отчета
    :param num_data:
    :param worksheet:
    :param violation_data: 
    :return: 
    """

    img_data = await read_json_file(violation_data)

    if not os.path.isfile(img_data['photo_full_name']):
        logger.error("photo not found")
        return False

    merged_cells = [
        f'C{start_photo_row + num_data}:E{start_photo_row + num_data}',
        f'G{start_photo_row + num_data}:H{start_photo_row + num_data}',
    ]

    for merged_cell in merged_cells:
        worksheet.merge_cells(merged_cell)

    photographic_materials_alignment = [
        f'C{start_photo_row + num_data}:H{start_photo_row + num_data}',
    ]

    for item, cell_range in enumerate(photographic_materials_alignment, start=1):
        await set_mip_alignment(worksheet, cell_range, horizontal='left', vertical='center')

    photographic_row_dimensions = [
        [f"{start_photo_row + num_data}", 300.0],
    ]

    for item in photographic_row_dimensions:
        worksheet.row_dimensions[int(item[0])].height = float(item[1])

    for item, cell_range in enumerate(photographic_materials_alignment, start=1):
        await set_report_font(worksheet, cell_range=cell_range, font_size=14)

    img: Image = Image(img_data['photo_full_name'])

    img_params: dict = {
        "column": photo_column,
        "row": start_photo_row + num_data,
        "height": photo_height,
        "anchor": True,
        "scale": True,
    }

    img = await image_preparation(img, img_params)

    await insert_images(worksheet, img=img)

    if not img_data.get('description'):
        return False

    worksheet.cell(row=start_photo_row + num_data, column=description_column, value=str(img_data['description']))

    for item, cell_border_range in enumerate(photographic_materials_alignment, start=1):
        await set_range_border(worksheet, cell_range=cell_border_range)

    worksheet.print_area = f'$A$1:$I${start_photo_row + num_data + 1}'

    return True
